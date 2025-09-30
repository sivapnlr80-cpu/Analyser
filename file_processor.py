import os
import re
import zipfile
import tempfile
import shutil
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime


class FileProcessor:
    """Main class for processing ZIP files containing schedules and annexures"""
    
    def __init__(self):
        self.temp_dir = None
        self.extracted_files = []
        # Schedule 2 to Schedule 21 (omitting Schedule 1 and Schedule 22)
        self.expected_schedules = [f"Schedule {i}" for i in range(2, 22)]
        # Annexure 2 to Annexure 12 (omitting Annexure 1)
        self.expected_annexures = [f"Annexure {i}" for i in range(2, 13)]
        # Files to omit from processing
        self.omitted_files = ["Schedule 1", "Schedule 22", "Annexure 1"]
        self.file_data = []
        self.missing_files = []
        self.society_name = ""
        self.society_number = ""
        
    def upload_zip_file(self, zip_path: str) -> bool:
        """Upload and validate ZIP file"""
        if not os.path.exists(zip_path):
            raise FileNotFoundError(f"ZIP file not found: {zip_path}")
        
        if not zipfile.is_zipfile(zip_path):
            raise ValueError("Selected file is not a valid ZIP file")
        
        return True
    
    def extract_zip(self, zip_path: str) -> str:
        """Extract ZIP file to temporary directory"""
        # Create temporary directory
        self.temp_dir = tempfile.mkdtemp(prefix="file_processor_")
        
        # Extract ZIP file
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(self.temp_dir)
        
        # Get all extracted files
        self.extracted_files = []
        for root, dirs, files in os.walk(self.temp_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    self.extracted_files.append(os.path.join(root, file))
        
        return self.temp_dir
    
    def identify_filenames(self) -> List[str]:
        """Identify all PDF filenames in extracted directory"""
        filenames = [os.path.basename(f) for f in self.extracted_files]
        return filenames
    
    def _should_omit_file(self, filename: str) -> bool:
        """Check if file should be omitted (Schedule 1, Schedule 22, Annexure 1)"""
        filename_upper = filename.upper()
        
        # Check for Schedule 1
        if re.search(r'SCHEDULE[\s\-_]*1(?!\d)', filename_upper):
            return True
        
        # Check for Schedule 22
        if re.search(r'SCHEDULE[\s\-_]*22(?!\d)', filename_upper):
            return True
        
        # Check for Annexure 1
        if re.search(r'ANNEXURE[\s\-_]*1(?!\d)', filename_upper) or \
           re.search(r'ANNEX[\s\-_]*1(?!\d)', filename_upper) or \
           re.search(r'ANX[\s\-_]*1(?!\d)', filename_upper):
            return True
        
        return False
    
    def _identify_schedule_number(self, filename_upper: str) -> Optional[int]:
        """Identify schedule number from filename (2-21, omitting 1 and 22)"""
        # Try different patterns to extract schedule number
        schedule_patterns = [
            r'SCHEDULE[\s\-_]*(\d+)',
            r'SCH[\s\-_]*(\d+)',
            r'SCHED[\s\-_]*(\d+)'
        ]
        
        for pattern in schedule_patterns:
            match = re.search(pattern, filename_upper)
            if match:
                num = int(match.group(1))
                # Only accept Schedule 2 to 21 (omit 1 and 22)
                if 2 <= num <= 21:
                    return num
        return None
    
    def _identify_annexure_number(self, filename_upper: str) -> Optional[int]:
        """Identify annexure number from filename (2-12, omitting 1)"""
        # Try different patterns to extract annexure number
        annexure_patterns = [
            r'ANNEXURE[\s\-_]*(\d+)',
            r'ANNEX[\s\-_]*(\d+)',
            r'ANX[\s\-_]*(\d+)'
        ]
        
        for pattern in annexure_patterns:
            match = re.search(pattern, filename_upper)
            if match:
                num = int(match.group(1))
                # Only accept Annexure 2 to 12 (omit 1)
                if 2 <= num <= 12:
                    return num
        return None
    
    def verify_schedules_annexures(self) -> Tuple[List[str], List[str]]:
        """Verify presence of required schedules (1-22) and annexures (1-12)"""
        found_schedules = set()
        found_annexures = set()
        
        for file_path in self.extracted_files:
            filename = os.path.basename(file_path)
            filename_upper = filename.upper()
            
            # Check for schedules (1 to 22)
            schedule_num = self._identify_schedule_number(filename_upper)
            if schedule_num:
                found_schedules.add(f"Schedule {schedule_num}")
            
            # Check for annexures (1 to 12)
            annexure_num = self._identify_annexure_number(filename_upper)
            if annexure_num:
                found_annexures.add(f"Annexure {annexure_num}")
        
        # Find missing files
        missing_schedules = [s for s in self.expected_schedules if s not in found_schedules]
        missing_annexures = [a for a in self.expected_annexures if a not in found_annexures]
        
        self.missing_files = missing_schedules + missing_annexures
        
        return missing_schedules, missing_annexures
    
    def extract_totals_from_pdf(self, pdf_path: str) -> Dict:
        """Extract Grand Total/Total values from each page of PDF"""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                num_pages = len(pdf_reader.pages)
                
                # Extract totals from each page
                page_totals = []
                for page_num in range(num_pages):
                    page_text = pdf_reader.pages[page_num].extract_text()
                    totals = self._parse_totals(page_text)
                    
                    if totals:  # Only add if totals were found
                        page_totals.append({
                            'page': page_num + 1,
                            'opening_balance': totals.get('opening_balance', 'N/A'),
                            'debit': totals.get('debit', 'N/A'),
                            'credit': totals.get('credit', 'N/A'),
                            'closing_balance': totals.get('closing_balance', 'N/A')
                        })
                
                return {
                    'filename': os.path.basename(pdf_path),
                    'total_pages': num_pages,
                    'page_totals': page_totals
                }
        except Exception as e:
            return {
                'filename': os.path.basename(pdf_path),
                'total_pages': 'Error',
                'page_totals': [],
                'error': str(e)
            }
    
    def _parse_totals(self, text: str) -> Dict[str, str]:
        """Parse total values from text"""
        totals = {}
        
        # Convert text to uppercase for case-insensitive matching
        text_upper = text.upper()
        lines = text.split('\n')
        
        # Look for Grand Total or Total lines
        for i, line in enumerate(lines):
            line_upper = line.upper()
            
            # Check for Grand Total or Total (but not subtotal)
            if ('GRAND TOTAL' in line_upper or 
                (('TOTAL' in line_upper) and ('SUBTOTAL' not in line_upper))):
                
                # Extract numbers from this line and possibly next lines
                numbers = re.findall(r'[\d,]+\.?\d*', line)
                
                # Clean numbers (remove commas)
                numbers = [n.replace(',', '') for n in numbers]
                
                # Try to identify which number is which
                if len(numbers) >= 4:
                    totals['opening_balance'] = numbers[0]
                    totals['debit'] = numbers[1]
                    totals['credit'] = numbers[2]
                    totals['closing_balance'] = numbers[3]
                elif len(numbers) >= 3:
                    # If we have 3 numbers, assume debit, credit, closing
                    totals['debit'] = numbers[0]
                    totals['credit'] = numbers[1]
                    totals['closing_balance'] = numbers[2]
                elif len(numbers) >= 2:
                    # If we have at least 2 numbers, assume they are debit and credit
                    totals['debit'] = numbers[0]
                    totals['credit'] = numbers[1]
                elif len(numbers) == 1:
                    # Single number might be closing balance
                    totals['closing_balance'] = numbers[0]
                
                break
        
        return totals
    
    def process_all_files(self):
        """Process all extracted PDF files (omitting Schedule 1, 22 and Annexure 1)"""
        self.file_data = []
        
        for pdf_path in self.extracted_files:
            filename = os.path.basename(pdf_path)
            
            # Skip files that should be omitted
            if self._should_omit_file(filename):
                continue
            
            data = self.extract_totals_from_pdf(pdf_path)
            self.file_data.append(data)
        
        # Sort by filename
        self.file_data.sort(key=lambda x: x['filename'])
    
    def generate_excel_report(self, output_path: str, society_name: str = "", society_number: str = ""):
        """Generate Excel report with findings including page-by-page totals"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Index"
        
        # Define styles
        header_font = Font(name='Arial', size=16, bold=True)
        title_font = Font(name='Arial', size=12, bold=True)
        table_header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
        missing_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        alt_row_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Top Heading - Society Name
        if society_name:
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = society_name.upper()
            cell.font = header_font
            cell.alignment = center_align
            current_row += 2
        
        # Title and Date
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"File Processing Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = title_font
        cell.alignment = center_align
        current_row += 2
        
        # Missing Files Section
        if self.missing_files:
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = "Missing Files"
            cell.font = title_font
            cell.fill = missing_fill
            cell.alignment = center_align
            current_row += 1
            
            for missing in self.missing_files:
                ws.merge_cells(f'A{current_row}:E{current_row}')
                cell = ws[f'A{current_row}']
                cell.value = missing
                cell.font = normal_font
                cell.alignment = center_align
                cell.border = thin_border
                current_row += 1
            
            current_row += 1
        else:
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = "✓ All required files are present"
            cell.font = title_font
            cell.alignment = center_align
            current_row += 2
        
        # File Details Section
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "File Details with Page-by-Page Totals"
        cell.font = title_font
        cell.alignment = center_align
        current_row += 2
        
        # Process each file
        for file_info in self.file_data:
            if file_info.get('page_totals') and len(file_info['page_totals']) > 0:
                # File header
                ws.merge_cells(f'A{current_row}:E{current_row}')
                cell = ws[f'A{current_row}']
                pages_with_totals = len(file_info['page_totals'])
                cell.value = f"File: {file_info['filename']} | Total Pages: {file_info['total_pages']} | Pages with Totals: {pages_with_totals}"
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.alignment = left_align
                current_row += 1
                
                # Table headers
                headers = ['Page', 'Opening Balance', 'Debit', 'Credit', 'Closing Balance']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = header
                    cell.font = table_header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1
                
                # Table data
                for idx, page_total in enumerate(file_info['page_totals']):
                    row_data = [
                        page_total['page'],
                        page_total['opening_balance'],
                        page_total['debit'],
                        page_total['credit'],
                        page_total['closing_balance']
                    ]
                    
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.value = value
                        cell.font = normal_font
                        cell.alignment = center_align
                        cell.border = thin_border
                        
                        # Alternate row colors
                        if idx % 2 == 1:
                            cell.fill = alt_row_fill
                    
                    current_row += 1
                
                current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        
        # Save workbook
        wb.save(output_path)
    
    def cleanup(self):
        """Clean up temporary directory"""
        if self.temp_dir and os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)


class FileProcessorGUI:
    """GUI for the file processor application"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("File Processor - Schedule & Annexure Verifier")
        self.root.geometry("850x650")
        self.root.resizable(True, True)
        
        self.processor = FileProcessor()
        self.zip_path = None
        self.society_name = ""
        self.society_number = ""
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Schedule & Annexure File Processor", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, pady=10)
        
        # Society Information section
        society_frame = ttk.LabelFrame(main_frame, text="Step 1: Society Information", padding="10")
        society_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=10)
        society_frame.columnconfigure(1, weight=1)
        
        ttk.Label(society_frame, text="Society Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.society_name_entry = ttk.Entry(society_frame, width=50)
        self.society_name_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        
        ttk.Label(society_frame, text="Society Number:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.society_number_entry = ttk.Entry(society_frame, width=50)
        self.society_number_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        
        # Upload section
        upload_frame = ttk.LabelFrame(main_frame, text="Step 2: Upload ZIP File", padding="10")
        upload_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=10)
        upload_frame.columnconfigure(1, weight=1)
        
        self.file_label = ttk.Label(upload_frame, text="No file selected")
        self.file_label.grid(row=0, column=0, columnspan=2, pady=5)
        
        upload_btn = ttk.Button(upload_frame, text="Select ZIP File", command=self.select_zip_file)
        upload_btn.grid(row=1, column=0, padx=5, pady=5)
        
        process_btn = ttk.Button(upload_frame, text="Process Files", command=self.process_files)
        process_btn.grid(row=1, column=1, padx=5, pady=5)
        
        # Output section
        output_frame = ttk.LabelFrame(main_frame, text="Processing Log", padding="10")
        output_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        output_frame.columnconfigure(0, weight=1)
        output_frame.rowconfigure(0, weight=1)
        
        # Text widget with scrollbar
        self.output_text = tk.Text(output_frame, wrap=tk.WORD, height=20)
        scrollbar = ttk.Scrollbar(output_frame, orient=tk.VERTICAL, command=self.output_text.yview)
        self.output_text.configure(yscrollcommand=scrollbar.set)
        
        self.output_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Download section
        download_frame = ttk.Frame(main_frame)
        download_frame.grid(row=4, column=0, pady=10)
        
        self.download_btn = ttk.Button(download_frame, text="Download Excel Report", 
                                       command=self.download_report, state=tk.DISABLED)
        self.download_btn.pack()
        
        self.report_path = None
    
    def select_zip_file(self):
        """Open file dialog to select ZIP file"""
        file_path = filedialog.askopenfilename(
            title="Select ZIP File",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        
        if file_path:
            self.zip_path = file_path
            self.file_label.config(text=f"Selected: {os.path.basename(file_path)}")
            self.log_message(f"ZIP file selected: {file_path}\n")
    
    def process_files(self):
        """Process the selected ZIP file"""
        if not self.zip_path:
            messagebox.showerror("Error", "Please select a ZIP file first!")
            return
        
        # Get society information
        self.society_name = self.society_name_entry.get().strip()
        self.society_number = self.society_number_entry.get().strip()
        
        self.output_text.delete(1.0, tk.END)
        self.log_message("Starting file processing...\n\n")
        
        # Log society information
        if self.society_name:
            self.log_message(f"Society Name: {self.society_name}\n")
        if self.society_number:
            self.log_message(f"Society Number: {self.society_number}\n")
        if self.society_name or self.society_number:
            self.log_message("\n")
        
        try:
            # Step 1: Validate ZIP file
            self.log_message("Step 1: Validating ZIP file...\n")
            self.processor.upload_zip_file(self.zip_path)
            self.log_message("✓ ZIP file is valid\n\n")
            
            # Step 2: Extract ZIP file
            self.log_message("Step 2: Extracting ZIP file...\n")
            temp_dir = self.processor.extract_zip(self.zip_path)
            self.log_message(f"✓ Files extracted to: {temp_dir}\n\n")
            
            # Step 3: Identify filenames
            self.log_message("Step 3: Identifying PDF files...\n")
            filenames = self.processor.identify_filenames()
            self.log_message(f"✓ Found {len(filenames)} PDF files\n")
            
            # Separate omitted files from processable files
            omitted = []
            processable = []
            for filename in filenames:
                if self.processor._should_omit_file(filename):
                    omitted.append(filename)
                else:
                    processable.append(filename)
            
            if processable:
                self.log_message(f"\nFiles to process: {len(processable)}\n")
                for filename in processable:
                    self.log_message(f"  - {filename}\n")
            
            if omitted:
                self.log_message(f"\n⚠ Omitted files (Schedule 1, 22, Annexure 1): {len(omitted)}\n", "warning")
                for filename in omitted:
                    self.log_message(f"  - {filename}\n", "warning")
            
            self.log_message("\n")
            
            # Step 4: Verify schedules and annexures
            self.log_message("Step 4: Verifying schedules and annexures...\n")
            missing_schedules, missing_annexures = self.processor.verify_schedules_annexures()
            
            if missing_schedules or missing_annexures:
                self.log_message("⚠ Missing files detected:\n", "warning")
                if missing_schedules:
                    self.log_message(f"  Missing Schedules: {', '.join(missing_schedules)}\n", "warning")
                if missing_annexures:
                    self.log_message(f"  Missing Annexures: {', '.join(missing_annexures)}\n", "warning")
            else:
                self.log_message("✓ All required schedules and annexures are present\n")
            self.log_message("\n")
            
            # Step 5: Process all files and extract totals from each page
            self.log_message("Step 5: Extracting totals from each page of PDF files...\n")
            self.processor.process_all_files()
            
            # Log summary of extracted data
            total_pages_with_totals = sum(len(f.get('page_totals', [])) for f in self.processor.file_data)
            self.log_message(f"✓ Processed {len(self.processor.file_data)} files\n")
            self.log_message(f"✓ Found totals on {total_pages_with_totals} pages\n\n")
            
            # Step 6: Generate Excel report
            self.log_message("Step 6: Generating Excel report...\n")
            self.report_path = os.path.join(tempfile.gettempdir(), 
                                           f"Audit_Index_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            self.processor.generate_excel_report(self.report_path, self.society_name, self.society_number)
            self.log_message(f"✓ Report generated: {self.report_path}\n\n")
            
            self.log_message("=" * 50 + "\n")
            self.log_message("Processing completed successfully!\n", "success")
            self.log_message("Click 'Download Excel Report' to save the report.\n", "success")
            
            self.download_btn.config(state=tk.NORMAL)
            
        except Exception as e:
            self.log_message(f"\n✗ Error: {str(e)}\n", "error")
            messagebox.showerror("Processing Error", f"An error occurred:\n{str(e)}")
    
    def download_report(self):
        """Save the generated report to user-selected location"""
        if not self.report_path:
            messagebox.showerror("Error", "No report available to download!")
            return
        
        save_path = filedialog.asksaveasfilename(
            title="Save Audit Index Report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Audit_Index_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if save_path:
            try:
                shutil.copy(self.report_path, save_path)
                messagebox.showinfo("Success", f"Report saved successfully to:\n{save_path}")
                self.log_message(f"\n✓ Report downloaded to: {save_path}\n", "success")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save report:\n{str(e)}")
    
    def log_message(self, message, tag=None):
        """Add message to output text widget"""
        self.output_text.insert(tk.END, message)
        if tag == "error":
            # Color the last inserted text red
            start_idx = self.output_text.index(f"end-{len(message)}c")
            self.output_text.tag_add("error", start_idx, tk.END)
            self.output_text.tag_config("error", foreground="red")
        elif tag == "success":
            start_idx = self.output_text.index(f"end-{len(message)}c")
            self.output_text.tag_add("success", start_idx, tk.END)
            self.output_text.tag_config("success", foreground="green")
        elif tag == "warning":
            start_idx = self.output_text.index(f"end-{len(message)}c")
            self.output_text.tag_add("warning", start_idx, tk.END)
            self.output_text.tag_config("warning", foreground="orange")
        
        self.output_text.see(tk.END)
        self.root.update()
    
    def on_closing(self):
        """Handle window closing"""
        self.processor.cleanup()
        self.root.destroy()


def main():
    """Main entry point"""
    root = tk.Tk()
    app = FileProcessorGUI(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
